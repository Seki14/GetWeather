# Tool name         :GetProgressChart(For Distribute)
# Module name       :GetProgressChart.py
# Detail            :This's the script to make Progress Chart.
# Implementer       :R.Ishikawa
# Version           :1.2
# Last update       :2020/06/16

# HISTORY
#1 Create New                                  Ver.1.0  R.I  2020/05/27
#2 Add procedure to be easy to use in Win10    Ver.1.1  R.I  2020/06/14
#3 Add 2 columns "No." and "担当者"            Ver.1.2  R.I  2020/06/16

import openpyxl.utils
import os

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule


wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '進捗管理表'


def cellwidth(start, end, wide):
    for length in range(start, end):
        alpha = sheet.cell(row=1, column=length).coordinate[:-1]
        sheet.column_dimensions['{}'.format(alpha)].width = wide


cellwidth(1, 8, 8.38)   # A-G列セルのサイズ設定
cellwidth(8, 9, 40)     # H列セルのサイズ設定
cellwidth(2, 3, 17.38)  # B列セルのサイズ設定
cellwidth(9, 10, 0.38)   # I列セルのサイズ設定
cellwidth(10, 336, 4.88) # J-LW列セルのサイズ設定

for rows in range(1, 100):
    sheet.row_dimensions[rows].height = 19

side1 = Side(style='thin', color='000000')
side2 = Side(style='double', color='000000')
side3 = Side(style='hair', color='000000')

for rows in sheet.iter_rows(min_row=1, min_col=1, max_row=100, max_col=9):
    for cell in rows:
        if cell.row == 5:
            cell.border = Border(right=side1, bottom=side2)
        else:
            cell.border = Border(right=side1)

for rows in sheet.iter_rows(min_row=1, min_col=9, max_row=100, max_col=335):
    for cell in rows:
        if cell.row == 5:
            cell.border = Border(right=side3, bottom=side2)
        else:
            cell.border = Border(top=side3, right=side3, bottom=side3, left=side3)

print("\n")
print("#######################################")
print("########## GetProgressChart ###########")
print("#######################################\n")
print("進捗管理表を自動生成するツールです。\n")

startYear = int(input('開始年を入力してください > '))  
startMonth = int(input('開始月を入力してください > ')) 
startDay = int(input('開始日を入力してください > '))   

sheet["A5"].value = "No."
sheet["B5"].value = "タスク"
sheet["C5"].value = "担当者"
sheet["D4"].value = "着手日"
sheet["D5"].value = "予定"
sheet["E5"].value = "実績"
sheet["F4"].value = "完了日"
sheet["F5"].value = "予定"
sheet["G5"].value = "実績"
sheet["H5"].value = "備考"
sheet["J2"].value = startYear
sheet["K2"].value = "年"
sheet["J3"].value = startMonth
sheet["K3"].value = "月"

sheet["J4"].number_format = "d"
sheet["J4"].value = '=DATE(J2, J3, {})'.format(startDay)
sheet["J5"].number_format = "aaa"
sheet["J5"].value = '=J4'


for rows in sheet.iter_rows(min_row=4, min_col=11, max_row=5, max_col=335):
    for cell in rows:
        if cell.row == 4:
            cell.number_format = 'd'
            cell.value = '={}+1'.format(sheet.cell(row=cell.row, column=cell.column-1).coordinate)
        else:
            cell.number_format = 'aaa'
            cell.value = '={}'.format(sheet.cell(row=cell.row-1, column=cell.column).coordinate)


for rows in sheet.iter_rows(min_row=3, min_col=12, max_row=3, max_col=335):
    for cell in rows:
        cell.number_format = "m"
        cell.value = '=IF(DAY({0})=1, {1},"")'.format(sheet.cell(column=cell.column, row=cell.row + 1).coordinate,
                                                      sheet.cell(column=cell.column, row=cell.row + 1).coordinate)

def colorMake(types, start, end):
    return PatternFill(fill_type=types, start_color=start, end_color=end)


grayfill = colorMake('solid', 'd3d3d3', 'd3d3d3')
planfill = colorMake('solid', '1e90ff', '1e90ff') #計画セル色。デフォルトは青。
actualfill = colorMake('solid', 'E4007f', 'E4007f') #実績セル色。デフォルトはマゼンダ。

for rows in sheet.iter_rows(min_row=6, min_col=2, max_row=100, max_col=8):
    for cell in rows:
        cell.number_format = "m/d"

# 完了日の実績が入力されたら、その行のタスクをグレー色にする
sheet.conditional_formatting.add('A6:H100', FormulaRule(formula=['NOT($G6="")'], stopIfTrue=True, fill=grayfill))
# 土曜日と日曜日をグレー色にする
sheet.conditional_formatting.add('J4:MF100', FormulaRule(formula=['OR(WEEKDAY(J$5)=1, WEEKDAY(J$5)=7)'], stopIfTrue=True, fill=grayfill))
# 着手日と完了日の実績が入力されたら、着手日～完了日のセルを実績色にする（デフォルト色：マゼンタ）
sheet.conditional_formatting.add('J6:MF100', FormulaRule(formula=['AND($E6<=J$4, $G6>=J$4)'], stopIfTrue=True, fill=actualfill))
# 着手日と完了日の予定が入力されたら、着手日～完了日のセルを予定色にする（デフォルト色：青）
sheet.conditional_formatting.add('J6:MF100', FormulaRule(formula=['AND($D6<=J$4, $F6>=J$4)'], stopIfTrue=True, fill=planfill))

# セル座標J6を固定にする
sheet.freeze_panes = 'J6'

# desktop_path = Desktopの絶対パス。実行環境によって変更する.
## Windowsの場合　'C:\Users\ユーザー名\Desktop'
## Macの場合 '/Users/ユーザ名/Desktop/' 
#desktop_path = '/Users/seki/Desktop/'

wb.save('進捗管理表.xlsx')
